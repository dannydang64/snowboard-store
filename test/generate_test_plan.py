#!/usr/bin/env python3
"""
Generate Test Plan Excel Script

This script converts structured JSON test execution results into a professionally formatted
Excel spreadsheet styled to resemble the Node Conversion.xlsx reference.
"""

import json
import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

def clean_string_for_excel(text):
    """Clean string to remove characters that Excel doesn't allow"""
    if text is None:
        return ""
    
    # Convert to string if it's not already
    text = str(text)
    
    # Replace problematic characters
    # Remove control characters
    text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', text)
    
    # Replace specific problematic characters and phrases
    text = text.replace('undefined', 'undefined value')
    
    # Replace characters that cause issues in Excel
    text = text.replace('\n', ' ')
    text = text.replace('\r', ' ')
    text = text.replace('\t', ' ')
    
    # Remove any characters that might cause issues with Excel
    text = text.replace('(', '[').replace(')', ']')
    text = text.replace('{', '[').replace('}', ']')
    
    # Replace problematic operators and expressions
    text = text.replace('===', 'equals')
    text = text.replace('==', 'equals')
    text = text.replace('!==', 'not equals')
    text = text.replace('!=', 'not equals')
    text = text.replace('expect[', 'expect ')
    text = text.replace('].toBe[', ' to be ')
    
    # Limit length to avoid Excel cell size issues
    if len(text) > 32000:  # Excel has a limit of 32,767 characters per cell
        text = text[:32000] + "..."
        
    return text

def extract_test_steps(test_name, failure_message=None):
    """Extract test steps from test name or create placeholder steps"""
    # Remove priority prefix (P0:, P1:, P2:)
    name = test_name.replace("P0: ", "").replace("P1: ", "").replace("P2: ", "")
    
    # Basic test steps based on test name
    if "add" in name.lower() and "product" in name.lower() and "cart" in name.lower():
        return "1. Navigate to product detail page\n2. Click 'Add to Cart' button\n3. Verify product is added to cart"
    elif "update" in name.lower() and "quantity" in name.lower():
        return "1. Add product to cart\n2. Navigate to cart page\n3. Update product quantity\n4. Verify quantity is updated"
    elif "remove" in name.lower() and "product" in name.lower():
        return "1. Add product to cart\n2. Navigate to cart page\n3. Click remove button\n4. Verify product is removed"
    elif "checkout" in name.lower():
        return "1. Add products to cart\n2. Navigate to checkout\n3. Fill shipping information\n4. Fill payment information\n5. Complete order\n6. Verify order confirmation"
    elif "api" in name.lower():
        if "crud" in name.lower():
            return "1. Send POST request to create resource\n2. Send GET request to retrieve resource\n3. Send PUT request to update resource\n4. Send DELETE request to remove resource"
        else:
            return "1. Send API request\n2. Verify response status code\n3. Validate response data"
    elif "search" in name.lower():
        return "1. Navigate to search page\n2. Enter search keyword\n3. Submit search\n4. Verify search results"
    elif "filter" in name.lower():
        return "1. Navigate to products page\n2. Select filter criteria\n3. Apply filter\n4. Verify filtered results"
    elif "persist" in name.lower():
        return "1. Add items to cart\n2. Refresh page\n3. Verify cart items are still present"
    else:
        # Generic steps
        return "1. Setup test environment\n2. Execute test actions\n3. Verify expected results"

def extract_test_data(test_name, category):
    """Extract test data based on test name and category"""
    if "product" in category:
        return "Product ID: SNOW-123\nProduct Name: Alpine Carver\nPrice: $599.99"
    elif "cart" in category:
        return "Product ID: SNOW-123\nQuantity: 2\nPrice: $599.99"
    elif "checkout" in category:
        return "Cart Items: 2 products\nSubtotal: $1,199.98\nShipping: $25.00\nTax: $98.00\nTotal: $1,322.98"
    elif "api" in category:
        return "Endpoint: /api/v1/products\nMethod: GET\nHeaders: {\"Content-Type\": \"application/json\"}"
    else:
        return ""  # Empty for other categories

def extract_expected_result(test_name):
    """Extract expected result from test name"""
    # Remove priority prefix
    name = test_name.replace("P0: ", "").replace("P1: ", "").replace("P2: ", "")
    
    # Convert "should X" to "X happens"
    if "should " in name.lower():
        parts = name.lower().split("should ")
        if len(parts) > 1:
            result = parts[1].strip()
            # Convert to present tense statement
            return f"System {result}"
    
    # For API tests
    if "api" in name.lower():
        if "return" in name.lower():
            return "API returns correct data with 200 status code"
        elif "crud" in name.lower():
            return "API supports all CRUD operations successfully"
        elif "error" in name.lower():
            return "API returns appropriate error codes and messages"
        else:
            return "API responds with expected data"
    
    # Category-specific expected results
    if "product" in name.lower():
        if "display" in name.lower() or "show" in name.lower():
            return "Products are displayed correctly with all details"
        elif "filter" in name.lower():
            return "Products are filtered according to selected criteria"
        elif "search" in name.lower():
            return "Search results display relevant products"
        elif "navigate" in name.lower():
            return "Navigation to product categories works correctly"
    elif "cart" in name.lower():
        if "add" in name.lower():
            return "Product is added to cart successfully"
        elif "update" in name.lower() or "quantity" in name.lower():
            return "Product quantity is updated correctly"
        elif "remove" in name.lower():
            return "Product is removed from cart successfully"
        elif "persist" in name.lower():
            return "Cart items persist after page refresh"
        elif "calculation" in name.lower():
            return "Cart totals are calculated correctly"
    elif "checkout" in name.lower():
        if "complete" in name.lower():
            return "Checkout process completes successfully"
        elif "order" in name.lower() and "summary" in name.lower():
            return "Order summary displays correct information"
        elif "confirmation" in name.lower():
            return "Order confirmation page shows correct order details"
        elif "reject" in name.lower() or "invalid" in name.lower():
            return "System rejects invalid payment information with appropriate error message"
    
    return f"Test completes successfully with expected outcome"

def extract_actual_result(test):
    """Extract actual result based on test status and failure messages"""
    if test["status"] == "passed":
        return "Test passed successfully"
    
    # For failed tests, extract meaningful information from failure messages
    if test["failureMessages"] and len(test["failureMessages"]) > 0:
        failure = test["failureMessages"][0]
        
        if "TypeError" in failure:
            if "Cannot read properties" in failure:
                return "Test failed: Cannot read properties of undefined object"
            elif "Cannot set properties" in failure:
                return "Test failed: Cannot set properties of undefined object"
        elif "AssertionError" in failure:
            return "Test failed: Assertion error - expected value not matching actual"
        elif "Timeout" in failure:
            return "Test failed: Operation timed out"
        
        # Extract first line of error message
        first_line = failure.split('\n')[0]
        return f"Test failed: {clean_string_for_excel(first_line)}"
    
    return "Test failed"

def determine_test_type(test_name):
    """Determine if test is positive or negative based on name"""
    negative_keywords = ["not", "invalid", "error", "fail", "reject", "empty"]
    for keyword in negative_keywords:
        if keyword in test_name.lower():
            return "Negative"
    return "Positive"

def create_test_plan_excel(json_file_path, output_excel_path):
    """Create formatted Excel test plan from JSON test results"""
    # Load JSON data
    with open(json_file_path, 'r') as file:
        data = json.load(file)
    
    # Extract test results
    passed_tests = data['tests']['passed']
    failed_tests = data['tests']['failed']
    all_tests = passed_tests + failed_tests
    
    # Create DataFrame for Excel
    excel_data = []
    
    # Process each test
    for i, test in enumerate(all_tests):
        test_name = test['name']
        category = test.get('category', 'other')
        
        # Determine test type (positive/negative)
        test_type = determine_test_type(test_name)
        
        # Extract test steps
        test_steps = extract_test_steps(test_name, test.get('failureMessages', []))
        
        # Extract test data
        test_data = extract_test_data(test_name, category)
        
        # Extract expected result
        expected_result = extract_expected_result(test_name)
        
        # Extract actual result
        actual_result = extract_actual_result(test)
        
        # Add to Excel data - clean all strings to avoid Excel issues
        excel_data.append({
            'SL No': i + 1,
            'Type': clean_string_for_excel(test_type),
            'Test Case Title': clean_string_for_excel(test_name.replace("P0: ", "").replace("P1: ", "").replace("P2: ", "")),
            'Test Steps': clean_string_for_excel(test_steps),
            'Test Data': clean_string_for_excel(test_data),
            'Expected Result': clean_string_for_excel(expected_result),
            'Actual Result': clean_string_for_excel(actual_result),
            'Comments': "",
            'Test Duration (Hours)': 2  # Default duration as requested
        })
    
    # Create DataFrame
    df = pd.DataFrame(excel_data)
    
    # Save to Excel
    df.to_excel(output_excel_path, index=False, sheet_name='Test Plan')
    
    # Format Excel file
    format_excel_file(output_excel_path)
    
    return output_excel_path

def format_excel_file(excel_path):
    """Format the Excel file with proper styling to match Node Conversion.xlsx"""
    wb = load_workbook(excel_path)
    ws = wb['Test Plan']
    
    # Define styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Light blue
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    prereq_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")  # Light orange
    
    pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green
    fail_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
    
    body_alignment = Alignment(vertical='center', wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title and pre-requisites section
    ws.insert_rows(1, 10)  # Insert rows at the top for title and pre-requisites
    
    # Title
    ws['A1'] = "Node Split Automation Functional Validation Scenarios"
    ws.merge_cells('A1:I1')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = header_fill
    
    # Pre-requisites
    ws['A3'] = "Pre-requisites:"
    ws['A3'].font = Font(bold=True)
    ws['A3'].fill = prereq_fill
    
    ws['A4'] = "1) Login to Storefront Test: https://snowboardstore.vercel.app/"
    ws['A4'].fill = prereq_fill
    
    ws['A5'] = "2) Navigate to the Store (assumes user has login credentials)"
    ws['A5'].fill = prereq_fill
    
    ws['A7'] = "Positive scenarios: 1) To check node fails at 0 if Loop OK fails, 2) Loop OK fails, 3) Loop Not fails"
    ws['A7'].fill = prereq_fill
    
    ws['A8'] = "Failure scenarios: 1) To check if node fails at 0 if Loop OK fails, 2) Loop OK fails, 3) Loop Not fails"
    ws['A8'].fill = prereq_fill
    
    ws['A10'] = "Note: Test Data which is provided is sample data for reference. Not all of the variables names will be selected from the list. Story driver will populate the values displayed on the day of UI"
    ws['A10'].fill = prereq_fill
    
    # Format pre-requisite section
    for row in range(1, 11):
        for col in range(1, 10):  # Columns A through I
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if row >= 3 and row <= 10:
                cell.fill = prereq_fill
    
    # Merge cells for pre-requisites
    for row in range(3, 11):
        ws.merge_cells(f'A{row}:I{row}')
        ws.row_dimensions[row].height = 20
    
    # Style header row (now at row 11)
    header_row = 11
    for col, cell in enumerate(ws[header_row], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Style data rows
    for row in range(header_row + 1, ws.max_row + 1):
        # Set row height
        ws.row_dimensions[row].height = 30
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = body_alignment
            cell.border = border
        
        # Color based on actual result
        actual_result_cell = ws.cell(row=row, column=7)  # Actual Result column
        if "passed" in str(actual_result_cell.value).lower():
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = pass_fill
        elif "failed" in str(actual_result_cell.value).lower():
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fail_fill
    
    # Auto-size columns
    dim_holder = DimensionHolder(worksheet=ws)
    
    # Set column widths
    column_widths = {
        1: 10,  # SL No
        2: 15,  # Type
        3: 30,  # Test Case Title
        4: 40,  # Test Steps
        5: 25,  # Test Data
        6: 30,  # Expected Result
        7: 30,  # Actual Result
        8: 20,  # Comments
        9: 15   # Test Duration
    }
    
    for col, width in column_widths.items():
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=width)
    
    ws.column_dimensions = dim_holder
    
    # Save workbook
    wb.save(excel_path)

if __name__ == "__main__":
    # File paths
    json_file_path = os.path.join(os.path.dirname(__file__), "reports", "test-results.json")
    output_excel_path = os.path.join(os.path.dirname(__file__), "reports", "test-plan-output.xlsx")
    
    # Generate Excel file
    excel_file = create_test_plan_excel(json_file_path, output_excel_path)
    print(f"Excel test plan generated successfully: {excel_file}")
